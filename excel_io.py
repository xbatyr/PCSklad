"""Экспорт в Excel и импорт со склада из Excel.

Файлы делаются настоящими .xlsx через openpyxl: у Excel нет проблем с кириллицей,
колонки уже нужной ширины, шапка закреплена, деньги отформатированы как числа.
"""

from datetime import datetime
from io import BytesIO
from typing import Iterable, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import CATEGORIES, Item, PC

# --- Оформление --------------------------------------------------------------

HEAD_FILL = PatternFill("solid", fgColor="1E293B")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="CBD5E1")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY = "# ##0"
DATE_FMT = "DD.MM.YYYY HH:MM"

TYPE_RU = {"component": "Комплектующая", "periphery": "Периферия", "service": "Услуга"}
TYPE_BACK = {v.lower(): k for k, v in TYPE_RU.items()}
STATUS_RU = {"in_stock": "На складе", "in_assembly": "В сборке", "sold": "Продано"}

CATEGORY_RU = {
    "CPU": "Процессор", "GPU": "Видеокарта", "RAM": "Оперативная память",
    "SSD": "SSD", "HDD": "Жёсткий диск", "Motherboard": "Мат. плата",
    "PSU": "Блок питания", "Case": "Корпус", "Cooler": "Охлаждение",
    "Monitor": "Монитор", "Mouse": "Мышь", "Keyboard": "Клавиатура",
    "Headset": "Наушники", "Speakers": "Колонки", "Microphone": "Микрофон",
    "Mousepad": "Коврик", "Repair": "Ремонт", "Maintenance": "Обслуживание",
    "Assembly": "Сборка", "Diagnostics": "Диагностика", "Software": "ПО и настройка",
    "Other": "Прочее",
}
CATEGORY_BACK = {v.lower(): k for k, v in CATEGORY_RU.items()}
CATEGORY_BACK.update({c.lower(): c for c in CATEGORIES})


def _header(ws, titles: list[str], widths: list[int], title: str) -> None:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Выгружено: {datetime.now():%d.%m.%Y %H:%M}")
    for col, (name, width) in enumerate(zip(titles, widths), start=1):
        cell = ws.cell(row=4, column=col, value=name)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:{get_column_letter(len(titles))}4"


def _row(ws, row_idx: int, values: list, money_cols: Iterable[int] = (), date_cols: Iterable[int] = ()) -> None:
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=value)
        cell.border = BORDER
        if col in money_cols:
            cell.number_format = MONEY
        elif col in date_cols:
            cell.number_format = DATE_FMT


# --- Экспорт -----------------------------------------------------------------


def build_workbook(pcs: list[PC], items: list[Item]) -> bytes:
    """Одна книга с тремя листами: сборки, склад, продажи."""
    wb = Workbook()

    # Лист 1 — готовые сборки
    ws = wb.active
    ws.title = "Сборки"
    _header(
        ws,
        ["№", "Название", "Статус", "Позиций", "Себестоимость", "Цена продажи", "Маржа", "Создан", "Продан"],
        [6, 42, 12, 9, 16, 16, 14, 18, 18],
        "ASTEX Stock — готовые ПК и сетапы",
    )
    for i, pc in enumerate(pcs, start=5):
        _row(ws, i, [
            pc.id, pc.name,
            "Продан" if pc.status == "sold" else "В наличии",
            len(pc.items), pc.cost_price, pc.sell_price, pc.margin,
            pc.created_at, pc.sold_at,
        ], money_cols=(5, 6, 7), date_cols=(8, 9))

    # Лист 2 — склад
    ws2 = wb.create_sheet("Склад")
    _header(
        ws2,
        ["№", "Тип", "Категория", "Наименование", "Закуп", "Розница",
         "Статус", "Расположение", "Откуда пришла", "Куплена", "Изменена", "Продана"],
        [6, 15, 18, 44, 13, 13, 12, 30, 30, 18, 18, 18],
        "ASTEX Stock — склад и позиции",
    )
    for i, it in enumerate(items, start=5):
        _row(ws2, i, [
            it.id, TYPE_RU.get(it.item_type, it.item_type),
            CATEGORY_RU.get(it.category, it.category), it.name,
            it.cost_price, it.retail_price,
            STATUS_RU.get(it.status, it.status),
            it.pc.name if it.pc else "Свободный склад",
            it.source_note or "",
            it.purchased_at or it.created_at, it.updated_at, it.sold_at,
        ], money_cols=(5, 6), date_cols=(10, 11, 12))

    # Лист 3 — продажи
    ws3 = wb.create_sheet("Продажи")
    _header(
        ws3,
        ["Что продано", "Тип", "Дата продажи", "Выручка", "Себестоимость", "Прибыль"],
        [46, 16, 18, 14, 16, 14],
        "ASTEX Stock — продажи",
    )
    sales: list[tuple] = []
    for pc in pcs:
        if pc.status == "sold":
            sales.append((pc.name, "Сборка", pc.sold_at, pc.sell_price, pc.cost_price, pc.margin))
    for it in items:
        if it.status == "sold" and it.pc_id is None:
            revenue = it.retail_price or 0
            kind = "Услуга" if it.item_type == "service" else "Розница"
            sales.append((it.name, kind, it.sold_at, revenue, it.cost_price, revenue - it.cost_price))
    sales.sort(key=lambda s: s[2] or datetime.min, reverse=True)

    for i, sale in enumerate(sales, start=5):
        _row(ws3, i, list(sale), money_cols=(4, 5, 6), date_cols=(3,))

    if sales:
        total = len(sales) + 5
        ws3.cell(row=total, column=1, value="ИТОГО").font = Font(bold=True)
        for col in (4, 5, 6):
            cell = ws3.cell(row=total, column=col,
                            value=f"=SUM({get_column_letter(col)}5:{get_column_letter(col)}{total - 1})")
            cell.font = Font(bold=True)
            cell.number_format = MONEY
            cell.border = BORDER

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_import_template() -> bytes:
    """Пустой шаблон для заполнения — чтобы заказчик не гадал, какие колонки нужны."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"
    _header(
        ws,
        ["Тип", "Категория", "Наименование", "Закуп", "Розница", "Откуда пришла", "Дата покупки"],
        [16, 20, 46, 13, 13, 32, 16],
        "Шаблон для загрузки товаров на склад",
    )
    примеры = [
        ["Комплектующая", "Видеокарта", "Palit RTX 4060 Dual 8GB", 175000, 205000, "Поставщик Алматы", datetime.now()],
        ["Периферия", "Мышь", "Logitech G102", 9500, 14000, "", datetime.now()],
    ]
    for i, row in enumerate(примеры, start=5):
        _row(ws, i, row, money_cols=(4, 5), date_cols=(7,))

    ws.cell(row=8, column=1,
            value="↑ Две строки выше — пример. Удалите их и впишите свои товары, начиная с 5-й строки.")
    ws.cell(row=9, column=1, value="Тип: Комплектующая / Периферия. Категорию пишите как в программе.")
    ws.cell(row=10, column=1, value="Обязательны только «Наименование» и «Закуп». Остальное можно оставить пустым.")
    for r in (8, 9, 10):
        ws.cell(row=r, column=1).font = Font(italic=True, color="64748B", size=10)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# --- Импорт ------------------------------------------------------------------


def _to_int(value) -> int:
    if value is None or value == "":
        return 0
    if isinstance(value, (int, float)):
        return int(round(value))
    text = str(value).replace(" ", "").replace("\xa0", "").replace("₸", "").replace(",", ".")
    try:
        return int(round(float(text)))
    except ValueError:
        raise ValueError(f"«{value}» — не похоже на число")


def _to_date(value) -> Optional[datetime]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    for fmt in ("%d.%m.%Y %H:%M", "%d.%m.%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return None


def parse_import(data: bytes) -> tuple[list[dict], list[str]]:
    """Читает загруженный файл. Возвращает (строки к добавлению, список замечаний)."""
    try:
        wb = load_workbook(BytesIO(data), data_only=True)
    except Exception:
        raise ValueError("Не удалось прочитать файл. Нужен .xlsx из Excel или Google Таблиц.")

    ws = wb["Склад"] if "Склад" in wb.sheetnames else wb.active

    # Ищем строку заголовков: первая, где встречается «наименование»
    header_row = None
    for row in range(1, min(ws.max_row, 20) + 1):
        values = [str(c.value or "").strip().lower() for c in ws[row]]
        if any("наименование" in v for v in values):
            header_row = row
            break
    if header_row is None:
        raise ValueError("В файле не найдена колонка «Наименование». Скачайте шаблон и заполните его.")

    headers = {}
    for cell in ws[header_row]:
        key = str(cell.value or "").strip().lower()
        if key:
            headers[key] = cell.column

    def col(*names) -> Optional[int]:
        for n in names:
            for key, idx in headers.items():
                if n in key:
                    return idx
        return None

    c_name = col("наименование", "название")
    c_cost = col("закуп", "себестоимость")
    c_type = col("тип")
    c_cat = col("категория")
    c_retail = col("розница", "розничная")
    c_src = col("откуда", "источник")
    c_date = col("дата покупки", "куплена")

    if c_name is None or c_cost is None:
        raise ValueError("Нужны колонки «Наименование» и «Закуп».")

    rows: list[dict] = []
    notes: list[str] = []

    for r in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=r, column=c_name).value
        if name is None or not str(name).strip():
            continue
        name = str(name).strip()
        if name.startswith("↑") or name.lower().startswith(("тип:", "обязательны")):
            continue

        try:
            cost = _to_int(ws.cell(row=r, column=c_cost).value)
        except ValueError as e:
            notes.append(f"Строка {r}: закуп — {e}. Пропущена.")
            continue

        raw_type = str(ws.cell(row=r, column=c_type).value or "").strip().lower() if c_type else ""
        item_type = TYPE_BACK.get(raw_type, "component" if not raw_type else None)
        if item_type is None:
            item_type = "periphery" if "периф" in raw_type else "component"
            notes.append(f"Строка {r}: тип «{raw_type}» не распознан, поставлен «{TYPE_RU[item_type]}».")

        raw_cat = str(ws.cell(row=r, column=c_cat).value or "").strip().lower() if c_cat else ""
        category = CATEGORY_BACK.get(raw_cat)
        if category is None:
            category = "Other"
            if raw_cat:
                notes.append(f"Строка {r}: категория «{raw_cat}» не распознана, поставлено «Прочее».")

        try:
            retail = _to_int(ws.cell(row=r, column=c_retail).value) if c_retail else 0
        except ValueError:
            retail = 0
            notes.append(f"Строка {r}: розничная цена не распознана, оставлена пустой.")

        rows.append({
            "item_type": item_type,
            "category": category,
            "name": name[:200],
            "cost_price": max(0, cost),
            "retail_price": retail or None,
            "source_note": (str(ws.cell(row=r, column=c_src).value).strip()[:500]
                            if c_src and ws.cell(row=r, column=c_src).value else None),
            "purchased_at": _to_date(ws.cell(row=r, column=c_date).value) if c_date else None,
        })

    if not rows:
        raise ValueError("В файле не нашлось ни одной строки с товаром.")
    return rows, notes
